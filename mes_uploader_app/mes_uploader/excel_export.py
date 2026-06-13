# -*- coding: utf-8 -*-
"""Tạo FILE MỚI (.xlsx) gom kết quả đo — mỗi lần đọc thêm 1 DÒNG, KÈM cột SN.

KHÔNG sao chép cả file gốc. Mỗi lần đọc chỉ lấy DÒNG CUỐI của file đo gốc (đọc
nhẹ, streaming) rồi APPEND vào file đích 1 dòng:

    SN, <nguyên dòng cuối của file gốc: Time, Judge, IspTime, Data01 … DataN>

GIỮ định dạng quan trọng của từng ô dòng cuối: MÀU CHỮ (vd ô đỏ = ngoài ngưỡng),
IN ĐẬM, NUMBER FORMAT (vd '0.000') và màu nền nếu có. Cột SN đặt ở ĐẦU. Dòng
tiêu đề (header) chỉ ghi 1 lần khi tạo file (in đậm như file gốc).

- Mỗi loại đầu 4X / 8X / 16X có 1 THƯ MỤC LƯU RIÊNG (xem resolve_output_dir).
- File: <output_dir>/<YYYYMMDD>/<tên file đo gốc>.xlsx
- Tự nhận diện file gốc theo NỘI DUNG: XLSX (bắt đầu 'PK', kể cả đuôi .csv) giữ
  được định dạng; CSV thật chỉ có giá trị (không có màu/định dạng).
- An toàn nhiều luồng (lock); 2 bên ghi 2 file khác nhau.

Module KHÔNG phụ thuộc Qt/PySide6 (để test headless và worker dùng được).
"""

import datetime
import io
import os
import sys
import threading

from .data_reader import csv_rows_from_bytes
from .i18n import tr

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:                          # cho phép import khi chưa cài openpyxl
    openpyxl = None
    Font = PatternFill = None

_lock = threading.Lock()

# Tên cột thêm vào (so với file gốc) — đặt ở ĐẦU.
SN_HEADER = "SN"


def default_output_dir():
    """Thư mục mặc định = 'excel_data' cạnh ứng dụng (exe hoặc thư mục run.py)."""
    if getattr(sys, "frozen", False):        # bản đóng gói PyInstaller
        base = os.path.dirname(sys.executable)
    else:                                    # chạy mã nguồn: .../mes_uploader_app
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, "excel_data")


def head_output_dir(excel_cfg, head_type):
    """Thư mục lưu RIÊNG theo loại đầu '4X'/'8X'/'16X'.

    Lùi về output_dir chung nếu thư mục riêng để trống. Trả về chuỗi đã strip
    (có thể rỗng nếu chưa cấu hình gì — bên gọi tự dùng mặc định).
    """
    per = {
        "4X": getattr(excel_cfg, "output_dir_4x", ""),
        "8X": getattr(excel_cfg, "output_dir_8x", ""),
        "16X": getattr(excel_cfg, "output_dir_16x", ""),
    }.get(head_type, "")
    shared = getattr(excel_cfg, "output_dir", "")
    return (per or "").strip() or (shared or "").strip()


def resolve_output_dir(cfg, head_type=None):
    """Thư mục lưu Excel từ cấu hình theo loại đầu; rỗng -> thư mục mặc định."""
    excel_cfg = getattr(cfg, "excel", None)
    d = head_output_dir(excel_cfg, head_type) if excel_cfg is not None else ""
    return d or default_output_dir()


def output_path(output_dir, source_file, when=None):
    """Đường dẫn file .xlsx đích: <output_dir>/<YYYYMMDD>/<tên file gốc>.xlsx."""
    when = when or datetime.datetime.now()
    base = os.path.splitext(os.path.basename(source_file or "data"))[0] or "data"
    return os.path.join(output_dir, when.strftime("%Y%m%d"), base + ".xlsx")


# ---------------------------------------------------------------------- #
#  Trích DÒNG CUỐI của file gốc + định dạng từng ô (đọc nhẹ, streaming)   #
# ---------------------------------------------------------------------- #
def _cell_style(cell, bold=None):
    """Rút gọn kiểu hiển thị 1 ô đủ để TÁI TẠO: giá trị + màu chữ + đậm/nghiêng
    + cỡ/tên font + number format + màu nền (nếu có)."""
    font = getattr(cell, "font", None)
    fill = getattr(cell, "fill", None)
    font_rgb = None
    if font is not None and font.color is not None:
        rgb = getattr(font.color, "rgb", None)
        if isinstance(rgb, str):
            font_rgb = rgb
    fill_rgb = fill_type = None
    if fill is not None and getattr(fill, "patternType", None):
        fill_type = fill.patternType
        fg = getattr(fill, "fgColor", None)
        rgb = getattr(fg, "rgb", None)
        if isinstance(rgb, str):
            fill_rgb = rgb
    return {
        "value": cell.value,
        "font_rgb": font_rgb,
        "bold": bool(font.bold) if (font and bold is None) else bool(bold),
        "italic": bool(font.italic) if font else False,
        "size": getattr(font, "size", None) if font else None,
        "name": getattr(font, "name", None) if font else None,
        "numfmt": getattr(cell, "number_format", "General"),
        "fill_rgb": fill_rgb,
        "fill_type": fill_type,
    }


def _plain_cell(value, bold=False):
    """Ô 'trơn' cho nguồn CSV (không có định dạng) — chỉ giữ giá trị (+ đậm header)."""
    return {"value": value, "font_rgb": None, "bold": bold, "italic": False,
            "size": None, "name": None, "numfmt": "General",
            "fill_rgb": None, "fill_type": None}


def extract_source(source_bytes):
    """Đọc NHẸ file gốc -> (header_cells, last_cells).

    XLSX (PK): mở read_only (streaming, KHÔNG nạp cả file vào RAM), chỉ giữ
    dòng tiêu đề + DÒNG DỮ LIỆU CUỐI kèm định dạng từng ô. CSV thật: tách dòng,
    lấy header + dòng cuối (không có định dạng). Trả về 2 list dict _cell_style.
    """
    if source_bytes[:2] == b"PK":            # XLSX (kể cả đuôi .csv)
        wb = openpyxl.load_workbook(io.BytesIO(source_bytes), read_only=True,
                                    data_only=True)
        try:
            ws = wb.active
            header = last = None
            for row in ws.iter_rows():
                if header is None:
                    header = row
                    continue
                if any(c.value not in (None, "") for c in row):
                    last = row
            header_cells = [_cell_style(c) for c in header] if header else []
            last_cells = [_cell_style(c) for c in last] if last else []
            return header_cells, last_cells
        finally:
            wb.close()

    rows = [r for r in csv_rows_from_bytes(source_bytes)
            if r and any(str(c).strip() for c in r)]
    if not rows:
        return [], []
    header_cells = [_plain_cell(v, bold=True) for v in rows[0]]
    last_cells = [_plain_cell(v) for v in rows[-1]]
    return header_cells, last_cells


# ---------------------------------------------------------------------- #
#  Ghi 1 ô / 1 dòng vào file đích (tái tạo định dạng)                     #
# ---------------------------------------------------------------------- #
def _apply_style(cell, info):
    if info.get("font_rgb") or info.get("bold") or info.get("italic"):
        cell.font = Font(name=info.get("name") or "Calibri",
                         size=info.get("size") or 11,
                         bold=bool(info.get("bold")),
                         italic=bool(info.get("italic")),
                         color=info.get("font_rgb"))
    nf = info.get("numfmt")
    if nf and nf != "General":
        cell.number_format = nf
    if info.get("fill_type") and info.get("fill_rgb"):
        cell.fill = PatternFill(start_color=info["fill_rgb"],
                                end_color=info["fill_rgb"],
                                fill_type=info["fill_type"])


def _write_cells(ws, row_idx, first_value, cells, first_bold=False):
    """Ghi 1 dòng: cột 1 = first_value (SN), các cột sau = cells (kèm định dạng)."""
    sn_cell = ws.cell(row=row_idx, column=1, value=first_value)
    if first_bold:
        sn_cell.font = Font(bold=True)
    for j, info in enumerate(cells, start=2):
        cell = ws.cell(row=row_idx, column=j, value=info.get("value"))
        _apply_style(cell, info)


def append_reading(out_path, sn, source_bytes):
    """Thêm 1 dòng (SN + dòng cuối file gốc, GIỮ định dạng) vào file đích.

    Tạo file + dòng tiêu đề (in đậm) khi chưa có. Trả về out_path. Cần openpyxl.
    """
    if openpyxl is None:
        raise RuntimeError(tr("Chưa cài thư viện 'openpyxl' để lưu Excel"))
    header_cells, last_cells = extract_source(source_bytes)
    if not last_cells:                       # file gốc chưa có dòng dữ liệu
        raise IOError(tr("File gốc chưa có dòng dữ liệu để lưu"))
    with _lock:
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        if os.path.exists(out_path):
            wb = openpyxl.load_workbook(out_path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data"
            _write_cells(ws, 1, SN_HEADER, header_cells, first_bold=True)
        _write_cells(ws, ws.max_row + 1, sn, last_cells)
        wb.save(out_path)
    return out_path
