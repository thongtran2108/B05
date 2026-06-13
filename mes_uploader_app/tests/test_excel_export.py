# -*- coding: utf-8 -*-
"""Kiểm thử tạo FILE MỚI gom kết quả đo (.xlsx) — append từng dòng, KÈM cột SN.

Trọng tâm theo yêu cầu mới:
- KHÔNG copy cả file gốc: tạo file mới, mỗi lần đọc chỉ lấy DÒNG CUỐI file gốc
  rồi APPEND 1 dòng (đọc nhẹ, không nạp cả file).
- GIỮ định dạng quan trọng của dòng cuối: MÀU CHỮ (ô đỏ), IN ĐẬM (header),
  NUMBER FORMAT ('0.000'). Cột SN ở ĐẦU.
- Gọi nhiều lần -> cộng dồn (header 1 lần + nhiều dòng).
- Mỗi loại đầu 4X / 8X / 16X chọn THƯ MỤC LƯU RIÊNG (resolve_output_dir).
- Worker (giả lập, 2 đầu 8X) bật Excel -> tạo file trong thư mục 8X, có cột SN.

Chạy:  python -m tests.test_excel_export
"""

import datetime
import glob
import io
import os
import sys
import tempfile
import time

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from openpyxl.styles import Font

from mes_uploader import excel_export as xe, mes_api
from mes_uploader.config import AppConfig, ExcelConfig, MaterialConfig, PathConfig
from mes_uploader.core.side_worker import SideWorker
from mes_uploader.hardware.plc_client import MockPlcClient

RED = "FFFF0000"
HEADERS = ["Time", "Judge", "IspTime", "Data01", "Data02"]


def _styled_source(values, red_cols=(2, 5)):
    """Dựng file đo gốc (.xlsx, bytes) giống cấu trúc thật: header in ĐẬM,
    vài ô dòng dữ liệu CHỮ ĐỎ (ngoài ngưỡng), cột Data dùng numfmt '0.000'."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    for j, h in enumerate(HEADERS, start=1):
        ws.cell(1, j, h).font = Font(bold=True)
    for j, v in enumerate(values, start=1):
        c = ws.cell(2, j, v)
        if j in red_cols:
            c.font = Font(color=RED)
        if j >= 4:                            # cột Data -> 3 chữ số thập phân
            c.number_format = "0.000"
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _csv_source(values):
    rows = [",".join(HEADERS), ",".join(str(v) for v in values)]
    return ("\n".join(rows) + "\n").encode("utf-8")


def _load(path):
    wb = openpyxl.load_workbook(path, data_only=False)
    return wb, wb.active


ROW1 = ["10:18:33", "OK", 1975, 24.2409262871991, 45.9486600992335]
ROW2 = ["10:18:40", "NG", 1949, 23.5012345678901, 22.0001112223334]


def test_output_path():
    print("== output_path ==")
    when = datetime.datetime(2026, 6, 9, 10, 18, 33)
    op = xe.output_path("/out", "/data/20260609/CCD1_NearStack.csv", when=when)
    assert op == os.path.join("/out", "20260609", "CCD1_NearStack.xlsx"), op
    print("  ", op)


def test_resolve_dir_per_head():
    print("\n== resolve_output_dir theo từng đầu ==")
    cfg = AppConfig()
    cfg.excel = ExcelConfig(enabled=True, output_dir="/common",
                            output_dir_4x="/d4", output_dir_16x="/d16")
    assert xe.resolve_output_dir(cfg, "4X") == "/d4"
    assert xe.resolve_output_dir(cfg, "16X") == "/d16"
    assert xe.resolve_output_dir(cfg, "8X") == "/common"   # chưa đặt -> chung
    cfg.excel = ExcelConfig(enabled=True)                  # trống hết -> mặc định
    assert xe.resolve_output_dir(cfg, "8X") == xe.default_output_dir()
    print("  4X=/d4  16X=/d16  8X->chung  (trống)->mặc định  OK")


def test_append_keeps_color_bold_numfmt():
    print("\n== append_reading: file MỚI, giữ màu chữ + đậm + numfmt, SN đầu ==")
    root = tempfile.mkdtemp(prefix="xl_")
    out = os.path.join(root, "CCD1.xlsx")

    xe.append_reading(out, "SN-1", _styled_source(ROW1))
    wb, ws = _load(out)
    try:
        # File MỚI: header (1) + 1 dòng dữ liệu = 2 dòng; SN thêm 1 cột ở đầu
        assert ws.max_row == 2, ws.max_row
        assert ws.max_column == len(HEADERS) + 1, ws.max_column
        # Tiêu đề: SN ở cột 1, IN ĐẬM; cột 2 = 'Time' cũng đậm
        assert ws.cell(1, 1).value == "SN" and ws.cell(1, 1).font.bold
        assert ws.cell(1, 2).value == "Time" and ws.cell(1, 2).font.bold
        # Dòng dữ liệu: SN + nguyên dòng gốc
        assert ws.cell(2, 1).value == "SN-1", ws.cell(2, 1).value
        assert ws.cell(2, 2).value == "10:18:33", ws.cell(2, 2).value
        # MÀU CHỮ ĐỎ giữ nguyên ở Judge (cột 3 trong đích) + Data02 (cột 6)
        assert RED in str(ws.cell(2, 3).font.color.rgb), ws.cell(2, 3).font.color.rgb
        assert RED in str(ws.cell(2, 6).font.color.rgb), ws.cell(2, 6).font.color.rgb
        # NUMBER FORMAT '0.000' giữ nguyên ở cột Data
        assert ws.cell(2, 5).number_format == "0.000", ws.cell(2, 5).number_format
        print("  SN-1 | Judge đỏ=%s | Data02 đỏ | numfmt=%s ✔"
              % (ws.cell(2, 3).font.color.rgb, ws.cell(2, 5).number_format))
    finally:
        wb.close()

    # Lần 2: CỘNG DỒN (không ghi đè) — header vẫn 1 dòng, thêm dòng SN-2
    print("\n== append_reading lần 2: cộng dồn (header 1 lần + 2 dòng) ==")
    xe.append_reading(out, "SN-2", _styled_source(ROW2, red_cols=(2,)))
    wb, ws = _load(out)
    try:
        assert ws.max_row == 3, ws.max_row
        assert ws.cell(2, 1).value == "SN-1", "dòng cũ phải còn"
        assert ws.cell(3, 1).value == "SN-2", ws.cell(3, 1).value
        assert ws.cell(3, 2).value == "10:18:40", ws.cell(3, 2).value
        print("  dòng2=SN-1 (giữ) | dòng3=SN-2 (mới)  OK")
    finally:
        wb.close()


def test_csv_source():
    print("\n== nguồn CSV thật -> vẫn ra file có cột SN (không có màu) ==")
    root = tempfile.mkdtemp(prefix="xl_csv_")
    out = os.path.join(root, "CCD2.xlsx")
    xe.append_reading(out, "SN-CSV", _csv_source(ROW1))
    wb, ws = _load(out)
    try:
        assert ws.cell(1, 1).value == "SN"
        assert ws.cell(2, 1).value == "SN-CSV", ws.cell(2, 1).value
        assert ws.cell(2, 2).value == "10:18:33", ws.cell(2, 2).value
        print("  OK (CSV -> SN + giá trị)")
    finally:
        wb.close()


def test_worker_uses_head_folder():
    print("\n== worker (2 đầu 8X) -> file trong THƯ MỤC 8X, có cột SN ==")
    root = tempfile.mkdtemp(prefix="xl_wk_")
    base = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                        "sample_data")
    dir_8x = os.path.join(root, "out_8x")
    dir_common = os.path.join(root, "out_common")

    cfg = AppConfig()
    cfg.simulation = True
    cfg.poll_interval_ms = 20
    cfg.paths = PathConfig(base_dir=base)
    cfg.paths.require_today = False
    cfg.materials = [MaterialConfig("ABC", heads_8x=2)]
    cfg.api.api_8x.url = "http://mes/8x/upload"
    cfg.excel.enabled = True
    cfg.excel.output_dir = dir_common          # chung (KHÔNG được dùng cho 8X)
    cfg.excel.output_dir_8x = dir_8x           # riêng cho 8X
    mes_api.post_payload = lambda u, p, **k: (True, 200, "OK")

    w = SideWorker("left", cfg, MockPlcClient(), lambda et, **d: None)
    w.start()
    w.arm(cfg.materials[0], "8X")
    time.sleep(0.1)
    w.submit_sn("SN-WK-7")
    time.sleep(0.2)
    for _ in range(2):
        w.simulate_trigger()
        time.sleep(0.3)
    time.sleep(0.6)
    w.stop()

    files_8x = glob.glob(os.path.join(dir_8x, "**", "*.xlsx"), recursive=True)
    files_common = glob.glob(os.path.join(dir_common, "**", "*.xlsx"), recursive=True)
    assert files_8x, "Phải tạo file trong thư mục 8X"
    assert not files_common, "KHÔNG được dùng thư mục chung khi đã đặt 8X"
    print("  file:", os.path.relpath(files_8x[0], dir_8x))

    wb, ws = _load(files_8x[0])
    try:
        assert ws.cell(1, 1).value == "SN", ws.cell(1, 1).value
        assert ws.cell(1, 2).value == "Time", ws.cell(1, 2).value
        data_rows = [r for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value]
        assert len(data_rows) == 2, "2 đầu -> 2 dòng, thực tế %d" % len(data_rows)
        assert all(ws.cell(r, 1).value == "SN-WK-7" for r in data_rows)
        print("  số cột=%d | 2 dòng SN-WK-7 ✔" % ws.max_column)
    finally:
        wb.close()


def test_disabled_no_file():
    print("\n== tắt Excel -> không tạo file ==")
    root = tempfile.mkdtemp(prefix="xl_off_")
    base = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                        "sample_data")
    cfg = AppConfig()
    cfg.simulation = True
    cfg.poll_interval_ms = 20
    cfg.paths = PathConfig(base_dir=base)
    cfg.paths.require_today = False
    cfg.materials = [MaterialConfig("ABC", heads_8x=2)]
    cfg.excel.enabled = False
    cfg.excel.output_dir_8x = root
    mes_api.post_payload = lambda u, p, **k: (True, 200, "OK")

    w = SideWorker("left", cfg, MockPlcClient(), lambda et, **d: None)
    w.start(); w.arm(cfg.materials[0], "8X"); time.sleep(0.1)
    w.submit_sn("SN-OFF"); time.sleep(0.2)
    for _ in range(2):
        w.simulate_trigger(); time.sleep(0.3)
    time.sleep(0.3); w.stop()
    assert glob.glob(os.path.join(root, "**", "*.xlsx"), recursive=True) == []
    print("  OK (không có file)")


def main():
    test_output_path()
    test_resolve_dir_per_head()
    test_append_keeps_color_bold_numfmt()
    test_csv_source()
    test_worker_uses_head_folder()
    test_disabled_no_file()
    print("\nTEST EXCEL-EXPORT PASS ✔")


if __name__ == "__main__":
    main()
